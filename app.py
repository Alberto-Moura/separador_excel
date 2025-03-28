import streamlit as st
import pandas as pd
import os
import json
import zipfile
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CONFIG_PATH = 'config_excel.json'

def carregar_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    
     # Definição de valores padrão caso não haja JSON
    config_padrao = {
        "cor_cabecalho": "0b480b",
        "cor_fonte_cabecalho": "FFFFFF",
        "tamanho_fonte_cabecalho": 23,
        "altura_linhas_cabecalho": 30,
        "alinhamento_vertical_cabecalho": "center",
        "alinhamento_horizontal_cabecalho": "center",
        "cor_fundo_tabela": "f9f0f0",
        "cor_texto_tabela": "000000",
        "tamanho_fonte_tabela": 20,
        "altura_linhas_tabela": 16,
        "alinhamento_vertical_texto": "center",
        "alinhamento_horizontal_texto": "center"
    }

    return config_padrao

def processar_planilha(uploaded_file, colunas_removidas, renomear_colunas, configuracao_excel):
    df = pd.read_excel(uploaded_file)
    df['Status'] = ''
    df['Nova Data'] = ''
    
    # Aplicar renomeação de colunas, se fornecido
    if renomear_colunas:
        df = df.rename(columns={k: v for k, v in renomear_colunas.items() if v})  # Ignora valores vazios
    
    # Normaliza os nomes para garantir que podemos localizar as colunas corretamente
    colunas_norm = {k: v for k, v in renomear_colunas.items() if v}  # Mapeia os novos nomes para os originais

    # Trata datas corretamente independentemente de renomeação
    data_emissao = colunas_norm.get('Data do documento', 'Data do documento')
    data_entrega = colunas_norm.get('Data de remessa', 'Data de remessa')

    if data_emissao in df.columns:
        df[data_emissao] = pd.to_datetime(df[data_emissao], errors='coerce').dt.strftime('%d/%m/%Y')
    if data_entrega in df.columns:
        df[data_entrega] = pd.to_datetime(df[data_entrega], errors='coerce').dt.strftime('%d/%m/%Y')

    if data_entrega in df.columns:
        df['Status'] = df[data_entrega].apply(lambda x: "Atrasado" if isinstance(x, str) and datetime.strptime(x, '%d/%m/%Y').date() < datetime.today().date() else "Dentro do prazo")
    
    # Remover colunas se necessário
    if colunas_removidas:
        df = df.drop(columns=colunas_removidas, errors='ignore')

    fornecedores_col = colunas_norm.get('Fornecedor/centro fornecedor', 'Fornecedor/centro fornecedor')
    if fornecedores_col not in df.columns:
        st.error("A coluna de fornecedores não foi encontrada na planilha!")
        return {}, 0, 0, 0

    fornecedores = df[fornecedores_col].dropna().unique()
    arquivos = {}

    for fornecedor in fornecedores:
        df_forn = df[df[fornecedores_col] == fornecedor]
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_forn.to_excel(writer, index=False, sheet_name='Planilha')
            #workbook = writer.book
            worksheet = writer.sheets['Planilha']
            
            worksheet.insert_rows(1)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_forn.columns))
            worksheet.cell(1, 1).value = f"Pedidos Pendentes - Fornecedor: {fornecedor}"
            worksheet.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(1, 1).font = Font(bold=True, size=16)
            worksheet.row_dimensions[1].height = 30
        
            # Aplicando estilos ao cabeçalho
            for col_num, col_name in enumerate(df_forn.columns, start=1):
                cell = worksheet.cell(row=2, column=col_num)
                cell.fill = PatternFill(start_color=configuracao_excel["cor_cabecalho"].replace("#", ""), fill_type="solid")
                cell.font = Font(bold=True, color=configuracao_excel["cor_fonte_cabecalho"].replace("#", ""), size=configuracao_excel["tamanho_fonte_cabecalho"])
                cell.alignment = Alignment(horizontal=configuracao_excel["alinhamento_horizontal_cabecalho"],
                                           vertical="center" if configuracao_excel["alinhamento_vertical_cabecalho"] == "middle" else configuracao_excel["alinhamento_vertical_cabecalho"])
                worksheet.row_dimensions[2].height = configuracao_excel["altura_linhas_cabecalho"]
            
            # Aplicando estilos à tabela
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=len(df_forn.columns)):
                for cell in row:
                    cell.fill = PatternFill(start_color=configuracao_excel["cor_fundo_tabela"].replace("#", ""), fill_type="solid")
                    cell.font = Font(color=configuracao_excel["cor_texto_tabela"].replace("#", ""), size=configuracao_excel["tamanho_fonte_tabela"])
                    cell.alignment = Alignment(horizontal=configuracao_excel["alinhamento_horizontal_texto"],
                                               vertical="center" if configuracao_excel["alinhamento_vertical_texto"] == "middle" else configuracao_excel["alinhamento_vertical_texto"])
            
            # Ajustando largura das colunas dinamicamente
            for col_num, col_name in enumerate(df_forn.columns, start=1):
                # Obtém o comprimento máximo da coluna considerando valores não nulos
                max_length = max(
                    df_forn[col_name].dropna().astype(str).map(len).max() if not df_forn[col_name].isna().all() else 0,
                    len(str(col_name))  # Inclui o cabeçalho
                )
                
                # Define a largura da coluna com uma margem extra
                worksheet.column_dimensions[get_column_letter(col_num)].width = max_length + 10

        output.seek(0)
        arquivos[f"{fornecedor}.xlsx"] = output

    pedidos_col = colunas_norm.get('Documento de compras', 'Documento de compras')
    pendente_col = colunas_norm.get('Qtd.pendente', 'Qtd.pendente')

    return arquivos, len(fornecedores), df[pedidos_col].nunique() if pedidos_col in df.columns else 0, df[pendente_col].sum() if pendente_col in df.columns else 0

st.title('Processamento de Pedidos')

uploaded_file = st.file_uploader('Faça o upload da Planilha', type=['xlsx'])

if uploaded_file:
    df = pd.read_excel(uploaded_file)

    st.divider()

    st.subheader('Resumo do Arquivo')
    col1, col2, col3 = st.columns(3)
    col1.metric('Fornecedores Únicos', df['Fornecedor/centro fornecedor'].nunique())
    col2.metric('Pedidos', df['Documento de compras'].nunique())
    col3.metric('Peças Pendentes', f"{df['Qtd.pendente'].sum():,.0f}".replace(',', '.'))

    st.divider()

    st.subheader('Excluir colunas')
    escolha_padrao = st.radio('Escolha configuração', ['Manter colunas padrão', 'Personalizar colunas'], horizontal=True,)
    colunas_padrao = ['Documento de compras', 'Item', 'Material', 'Valor da matriz', 'Texto breve', 'Qtd.divisão', 'Qtd.fornecida', 'Qtd.pendente', 'Data do documento', 'Data de remessa', 'Fornecedor/centro fornecedor', 'Centro']
    
    colunas_disponiveis = df.columns

    if escolha_padrao == 'Manter colunas padrão':
        colunas_removidas = st.multiselect('Selecione as colunas que deseja remover', colunas_disponiveis, 
                                       default=[col for col in colunas_disponiveis if col not in colunas_padrao])
    else:
        colunas_removidas = st.multiselect('Selecione as colunas que deseja remover', colunas_disponiveis)
    
    st.subheader('Renomear Colunas')

    renomear_colunas = {}
    coluna_par, coluna_meio, coluna_impar = st.columns([1, 0.05, 1])

    dividir_coluna = 0
    for i, coluna in enumerate(colunas_disponiveis):
        if coluna in colunas_removidas:
            continue
        if dividir_coluna % 2 == 0:
            with coluna_par:
                novo_nome = st.text_input(f'Renomear {coluna.upper()} para:', value=coluna)
                renomear_colunas[coluna] = novo_nome
        else:
            with coluna_impar:
                novo_nome = st.text_input(f'Renomear {coluna.upper()} para:', value=coluna)
                renomear_colunas[coluna] = novo_nome
        coluna_meio.markdown(f'<div style="border-left: 1px solid grey; height: {len(colunas_disponiveis)/2*2,5294}px;"></div>', unsafe_allow_html=True)
        dividir_coluna += 1

    st.divider()
    st.subheader('Estilização da Planilha')
    configuracao_excel = carregar_config()

    if configuracao_excel:
        st.success('Configuração localizada com sucesso!')
    else:
        st.warning('Nenhuma pré-configuração localizada. Faça o upload abaixo ou crie sua própria neste [link](conifg.py).')
        estilo = st.file_uploader('Caso tenha uma configuração pronta, Faça o upload do arquivo de configuração', type=['json'])

        if estilo is not None:
            configuracao_excel = json.load(estilo)
            with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
                json.dump(configuracao_excel, f, indent=4)
            st.success('Configuração carregada com sucesso!')
    
    st.write('Configuração atual:')

    # Criar tabela com cores representadas como quadrados
    styled_data = []
    for key, value in configuracao_excel.items():
        if "cor" in key:
            value = f'<div style="width: 30px; height: 20px; background-color: {value}; border: 1px solid #000;"></div>'
        styled_data.append((key, value))

    # Criar DataFrame estilizado
    df_config = pd.DataFrame(styled_data, columns=["Propriedade", "Valor"])

    st.write(df_config.to_html(escape=False), unsafe_allow_html=True)

    st.link_button('Ir para configuração de estilo', 'config.py',)

    st.divider()

    if st.button('Processar Arquivo', type='primary', key='processar_arquivos'):
        arquivos, qtd_fornecedores, qtd_pedidos, qtd_pecas = processar_planilha(uploaded_file, colunas_removidas, renomear_colunas, configuracao_excel)
        st.session_state.arquivos_processados = arquivos  # Armazena no session_state
        st.session_state.selecionados = list(arquivos.keys())  # Define padrão com todos os arquivos selecionados

# Exibe a seleção de fornecedores apenas se houver arquivos processados
if "arquivos_processados" in st.session_state and st.session_state.arquivos_processados and uploaded_file:
    nova_selecao = st.multiselect("Selecione os arquivos para download em massa",
                                  list(st.session_state.arquivos_processados.keys()),
                                  default=st.session_state.selecionados)

    # Atualiza a seleção apenas se houver mudanças
    if set(nova_selecao) != set(st.session_state.selecionados):
        st.session_state.selecionados = nova_selecao

    # Criar e oferecer o download do ZIP apenas se houver arquivos selecionados
    if st.session_state.selecionados:
        st.subheader('Download em Massa')
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for nome in st.session_state.selecionados:
                zipf.writestr(nome, st.session_state.arquivos_processados[nome].getvalue())
        zip_buffer.seek(0)
        st.download_button("Baixar Arquivos Selecionados", zip_buffer, "arquivos.zip", "application/zip", key='zip_download')
    
    # Botões de download individuais
    st.subheader('Download individual')
    for nome, buffer in st.session_state.arquivos_processados.items():
        st.download_button(
            label=f"Baixar {nome}",
            data=buffer,
            file_name=nome,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_{nome}"
        )


